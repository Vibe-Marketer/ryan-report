"""Ryan Report -- Desktop App

A simple desktop UI for running the Ryan report pipeline.
Double-click to launch, click a button to run.
"""
from __future__ import annotations

import json
import os
import platform
import sys
import threading
import time
from pathlib import Path
from typing import Any

import webview


# ---------------------------------------------------------------------------
# Path resolution -- works both in dev and when frozen by PyInstaller.
# ---------------------------------------------------------------------------

def _app_root() -> Path:
    """Return the project root (one level above app/)."""
    if getattr(sys, "frozen", False):
        # PyInstaller puts the exe in dist/; the repo is bundled inside _MEIPASS.
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent.parent


def _ui_path() -> str:
    return str(Path(__file__).resolve().parent / "ui" / "index.html")


APP_ROOT = _app_root()
EXECUTION = APP_ROOT / "execution"
STATE = APP_ROOT / "state"


def _user_config_dir() -> Path:
    system = platform.system()
    home = Path.home()
    if system == "Darwin":
        return home / "Library" / "Application Support" / "Catom"
    if system == "Windows":
        appdata = Path(os.environ.get("APPDATA", home / "AppData" / "Roaming"))
        return appdata / "Catom"
    return home / ".config" / "catom"


def _user_config_path() -> Path:
    return _user_config_dir() / "browser_config.json"


def _template_config_path() -> Path:
    example = EXECUTION / "browser_config.example.json"
    if example.exists():
        return example
    return EXECUTION / "browser_config.json"


def _legacy_bundled_config_paths() -> list[Path]:
    return [
        EXECUTION / "browser_config.json",
        APP_ROOT / "_internal" / "execution" / "browser_config.json",
        APP_ROOT / "app" / "_internal" / "execution" / "browser_config.json",
    ]


def _remove_legacy_bundled_configs() -> None:
    for path in _legacy_bundled_config_paths():
        try:
            if path.exists():
                path.unlink()
        except OSError:
            # Ignore cleanup failures. The app should still prefer user config.
            pass


# ---------------------------------------------------------------------------
# Pipeline API -- exposed to the webview JS layer.
# ---------------------------------------------------------------------------

class PipelineAPI:
    """Methods callable from the browser UI via window.pywebview.api.*"""

    def __init__(self, window: webview.Window | None = None):
        self._window = window
        self._running = False
        self._log_lines: list[str] = []
        _remove_legacy_bundled_configs()

    def set_window(self, window: webview.Window) -> None:
        self._window = window

    # -- Missing file bridge --

    def _request_missing_file(self, path: str) -> str:
        """Ask the UI what to do about a missing file. Blocks pipeline thread.
        Returns: a file path, 'new', or 'cancel'."""
        if not self._window:
            return "new"
        self._missing_file_result: str | None = None
        self._missing_file_event = threading.Event()
        import json as _json
        self._window.evaluate_js(
            f"showMissingFile({_json.dumps(path)}).then(r => window.pywebview.api.submit_missing_file(r))"
        )
        self._missing_file_event.wait(timeout=300)
        return (self._missing_file_result or "cancel").strip()

    def submit_missing_file(self, result: str) -> str:
        """Called from JS with a file path, 'new', or 'cancel'."""
        self._missing_file_result = result
        if hasattr(self, "_missing_file_event"):
            self._missing_file_event.set()
        return "ok"

    # -- Config helpers --

    def get_config_path(self) -> str:
        p = _user_config_path()
        p.parent.mkdir(parents=True, exist_ok=True)
        return str(p)

    def load_config(self) -> dict[str, Any]:
        user_path = Path(self.get_config_path())
        source_path = user_path if user_path.exists() else _template_config_path()
        if not source_path.exists():
            return {}
        with source_path.open("r") as f:
            cfg = json.load(f)
        # Expand ${HOME} etc.
        def _exp(o: Any) -> Any:
            if isinstance(o, str):
                return os.path.expandvars(o)
            if isinstance(o, dict):
                return {k: _exp(v) for k, v in o.items()}
            if isinstance(o, list):
                return [_exp(v) for v in o]
            return o
        return _exp(cfg)

    def save_config(self, cfg: dict) -> str:
        p = Path(self.get_config_path())
        p.parent.mkdir(parents=True, exist_ok=True)
        with p.open("w") as f:
            json.dump(cfg, f, indent=2)
        return "ok"

    def validate_config(self) -> dict[str, list[str]]:
        cfg = self.load_config()
        errors: list[str] = []
        warnings: list[str] = []

        browser = cfg.get("browser", {})
        auth = cfg.get("auth", {})
        downloads = cfg.get("downloads", {})

        exe_path = browser.get("executable_path", "")
        if not exe_path:
            errors.append("Browser executable path is missing.")
        elif not Path(exe_path).exists():
            errors.append(f"Browser executable does not exist: {exe_path}")

        user_data_dir = browser.get("user_data_dir", "")
        if not user_data_dir:
            errors.append("Browser user data directory is missing.")
        elif not Path(user_data_dir).exists():
            warnings.append(f"Browser user data directory does not exist yet: {user_data_dir}")

        if not auth.get("base_url"):
            errors.append("Axon base URL is missing.")
        if not auth.get("username"):
            errors.append("Axon username is missing.")
        if not auth.get("password"):
            errors.append("Axon password is missing.")

        download_dir = downloads.get("directory", "")
        if not download_dir:
            errors.append("Download directory is missing.")
        else:
            dl = Path(download_dir)
            try:
                dl.mkdir(parents=True, exist_ok=True)
            except Exception as exc:
                errors.append(f"Could not create download directory: {dl} ({exc})")

        historical = cfg.get("historical_ryan", "")
        if not historical:
            warnings.append("Historical Ryan CSV is not set -- all orders will be treated as new.")
        elif not Path(historical).exists():
            errors.append(f"Historical Ryan CSV does not exist: {historical}")

        reports = [r for r in cfg.get("reports", []) if r.get("enabled", True)]
        if not reports:
            errors.append("No enabled reports are configured.")

        return {"errors": errors, "warnings": warnings}

    # -- Auto-detection --

    def is_configured(self) -> bool:
        """Return True only when the user has saved a real config."""
        user_path = Path(self.get_config_path())
        if not user_path.exists():
            return False
        cfg = self.load_config()
        return bool(cfg.get("auth", {}).get("username"))

    def detect_browsers(self) -> list[dict[str, str]]:
        """Scan for supported Chromium-based browsers. Returns list of
        {name, path, user_data_dir} for each found browser."""
        import platform as plat

        browsers: list[dict[str, str]] = []
        is_mac = plat.system() == "Darwin"
        is_win = plat.system() == "Windows"
        is_linux = plat.system() == "Linux"
        home = Path.home()

        candidates: list[tuple[str, str, str]] = []
        if is_mac:
            candidates = [
                ("Comet",
                 "/Applications/Comet.app/Contents/MacOS/Comet",
                 str(home / "Library/Application Support/Comet")),
                ("Google Chrome",
                 "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
                 str(home / "Library/Application Support/Google/Chrome")),
                ("Chromium",
                 "/Applications/Chromium.app/Contents/MacOS/Chromium",
                 str(home / "Library/Application Support/Chromium")),
                ("Brave Browser",
                 "/Applications/Brave Browser.app/Contents/MacOS/Brave Browser",
                 str(home / "Library/Application Support/BraveSoftware/Brave-Browser")),
                ("Microsoft Edge",
                 "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
                 str(home / "Library/Application Support/Microsoft Edge")),
            ]
        elif is_win:
            localappdata = os.environ.get("LOCALAPPDATA", "")
            programfiles = os.environ.get("PROGRAMFILES", "C:\\Program Files")
            programfiles86 = os.environ.get("PROGRAMFILES(X86)", "C:\\Program Files (x86)")
            candidates = [
                ("Comet",
                 f"{localappdata}\\Programs\\Comet\\Comet.exe",
                 f"{localappdata}\\Comet\\User Data"),
                ("Google Chrome",
                 f"{programfiles}\\Google\\Chrome\\Application\\chrome.exe",
                 f"{localappdata}\\Google\\Chrome\\User Data"),
                ("Google Chrome (x86)",
                 f"{programfiles86}\\Google\\Chrome\\Application\\chrome.exe",
                 f"{localappdata}\\Google\\Chrome\\User Data"),
                ("Chromium",
                 f"{programfiles}\\Chromium\\Application\\chromium.exe",
                 f"{localappdata}\\Chromium\\User Data"),
                ("Chromium (x86)",
                 f"{programfiles86}\\Chromium\\Application\\chromium.exe",
                 f"{localappdata}\\Chromium\\User Data"),
                ("Brave Browser",
                 f"{programfiles}\\BraveSoftware\\Brave-Browser\\Application\\brave.exe",
                 f"{localappdata}\\BraveSoftware\\Brave-Browser\\User Data"),
                ("Microsoft Edge",
                 f"{programfiles}\\Microsoft\\Edge\\Application\\msedge.exe",
                 f"{localappdata}\\Microsoft\\Edge\\User Data"),
                ("Microsoft Edge (x86)",
                 f"{programfiles86}\\Microsoft\\Edge\\Application\\msedge.exe",
                 f"{localappdata}\\Microsoft\\Edge\\User Data"),
            ]
        elif is_linux:
            candidates = [
                ("Comet", "/usr/bin/comet", str(home / ".config" / "Comet")),
                ("Google Chrome", "/usr/bin/google-chrome", str(home / ".config" / "google-chrome")),
                ("Chromium", "/usr/bin/chromium", str(home / ".config" / "chromium")),
            ]

        for name, exe, user_data in candidates:
            if Path(exe).exists():
                # Chrome blocks CDP on its default data directory.
                # Use a dedicated Catom automation profile instead.
                if "Google Chrome" in name:
                    catom_profile = str(_user_config_dir() / "ChromeProfile")
                    Path(catom_profile).mkdir(parents=True, exist_ok=True)
                    user_data = catom_profile
                browsers.append({
                    "name": name,
                    "path": exe,
                    "user_data_dir": user_data,
                })

        return browsers

    def detect_profiles(self, user_data_dir: str) -> list[str]:
        """List available Chrome profile directories."""
        ud = Path(user_data_dir)
        if not ud.exists():
            return ["Default"]
        profiles = []
        if (ud / "Default").exists():
            profiles.append("Default")
        for p in sorted(ud.iterdir()):
            if p.is_dir() and p.name.startswith("Profile "):
                profiles.append(p.name)
        return profiles if profiles else ["Default"]

    def get_default_download_dir(self) -> str:
        return str(Path.home() / "Downloads" / "ryan-moves-and-tests")

    # -- File/folder browser --

    def pick_folder(self) -> str | None:
        result = self._window.create_file_dialog(
            webview.FOLDER_DIALOG, allow_multiple=False
        )
        if result and len(result) > 0:
            return result[0]
        return None

    def pick_file(self, title: str = "Select file") -> str | None:
        result = self._window.create_file_dialog(
            webview.OPEN_DIALOG,
            allow_multiple=False,
            file_types=("Spreadsheets (*.csv;*.xlsx;*.xls)", "CSV Files (*.csv)", "Excel Files (*.xlsx;*.xls)", "All Files (*.*)"),
        )
        if result and len(result) > 0:
            return result[0]
        return None

    def ensure_folder(self, path: str) -> str:
        """Create a folder if it doesn't exist. Returns 'ok'."""
        Path(path).mkdir(parents=True, exist_ok=True)
        return "ok"

    # -- Hours saved tracking --

    def get_hours_saved(self) -> float:
        """Return total hours saved from config."""
        cfg = self.load_config()
        return float(cfg.get("hours_saved", 0))

    def reset_hours_saved(self) -> str:
        """Reset the hours saved counter to 0."""
        cfg = self.load_config()
        cfg["hours_saved"] = 0
        self.save_config(cfg)
        return "ok"

    def _increment_hours_saved(self, hours: float = 3.0) -> float:
        """Add hours to the saved counter and return new total."""
        cfg = self.load_config()
        current = float(cfg.get("hours_saved", 0))
        new_total = current + hours
        cfg["hours_saved"] = new_total
        self.save_config(cfg)
        return new_total

    # -- Pipeline execution --

    def is_running(self) -> bool:
        return self._running

    def get_logs(self) -> list[str]:
        """Return and flush accumulated log lines."""
        lines = self._log_lines[:]
        self._log_lines.clear()
        return lines

    def run_pipeline(self, mode: str = "all") -> str:
        """Run the pipeline in a background thread. mode: all|download|build"""
        if self._running:
            return "already running"
        self._running = True
        self._log_lines.clear()
        t = threading.Thread(target=self._run, args=(mode,), daemon=True)
        t.start()
        return "started"

    def _log(self, msg: str) -> None:
        self._log_lines.append(msg)

    def _append_to_xlsx(self, historical_path: str, new_rows_csv: str) -> None:
        """Append new rows from the generated CSV to the xlsx version of the
        historical file. If the xlsx doesn't exist yet, create it."""
        try:
            import csv as csv_mod
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self._log("[WARN] openpyxl not installed -- skipping xlsx append.")
            return

        new_csv = Path(new_rows_csv)
        if not new_csv.exists() or new_csv.stat().st_size == 0:
            return

        # Read the new rows from the generated CSV.
        with new_csv.open("r", encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f)
            all_rows = list(reader)
        if len(all_rows) < 3:
            # First 2 rows are headers, need at least 1 data row.
            return
        headers = all_rows[:2]
        data_rows = all_rows[2:]

        # Determine the xlsx path -- same name/location as historical but .xlsx
        hist = Path(historical_path)
        xlsx_path = hist.with_suffix(".xlsx")

        if xlsx_path.exists():
            wb = load_workbook(xlsx_path)
            ws = wb.active
            # Append only data rows (headers already exist).
            for row in data_rows:
                ws.append(row)
            self._log(f"[OK] Appended {len(data_rows)} rows to {xlsx_path.name}")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ryan Moves"
            for row in headers + data_rows:
                ws.append(row)
            self._log(f"[OK] Created {xlsx_path.name} with {len(data_rows)} rows")

        wb.save(xlsx_path)

    def _latest_matching_file(self, directory: Path, prefix: str) -> Path | None:
        matches = sorted(
            directory.glob(f"{prefix}*.csv"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        return matches[0] if matches else None

    def _preflight_build_inputs(self, cfg: dict[str, Any]) -> list[str]:
        errors: list[str] = []
        dl_dir = Path(cfg.get("downloads", {}).get("directory", ""))
        if not dl_dir.exists():
            return [f"Download directory does not exist: {dl_dir}"]

        # Historical file is handled separately with the missing file dialog.
        # Don't block the build here -- just check for required source CSVs.
        required_prefixes = ["Order Master Report", "New RYAN"]
        for prefix in required_prefixes:
            if not self._latest_matching_file(dl_dir, prefix):
                errors.append(f"Required source CSV not found in downloads: {prefix}*.csv")

        return errors

    def _download_timeout_seconds(self, cfg: dict[str, Any]) -> int:
        auth = cfg.get("auth", {})
        reports = [r for r in cfg.get("reports", []) if r.get("enabled", True)]
        manual_login_timeout = int(auth.get("manual_login_timeout_seconds", 180))
        per_report_budget = 120
        launch_buffer = 180
        return max(900, manual_login_timeout + (len(reports) * per_report_budget) + launch_buffer)

    def _build_timeout_seconds(self) -> int:
        return 900

    def _run_command_streaming(
        self,
        cmd: list[str],
        timeout: int,
        cwd: Path | None = None,
    ) -> int:
        import subprocess

        process = subprocess.Popen(
            cmd,
            cwd=cwd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )

        start = time.time()
        assert process.stdout is not None
        try:
            for raw_line in iter(process.stdout.readline, ""):
                line = raw_line.rstrip("\n")
                if line:
                    self._log(line)
                if time.time() - start > timeout:
                    process.kill()
                    raise subprocess.TimeoutExpired(cmd, timeout)

            return process.wait(timeout=max(1, int(timeout - (time.time() - start))))
        except subprocess.TimeoutExpired:
            process.kill()
            raise
        finally:
            if process.stdout:
                process.stdout.close()

    def _run(self, mode: str) -> None:
        try:
            config = self.get_config_path()
            cfg = self.load_config()

            validation = self.validate_config()
            for warning in validation["warnings"]:
                self._log(f"[WARN] {warning}")
            if validation["errors"]:
                for error in validation["errors"]:
                    self._log(f"[ERROR] {error}")
                self._log("[ERROR] Preflight validation failed. Fix settings and try again.")
                return

            # Import the execution modules -- works in both dev and frozen modes.
            sys.path.insert(0, str(EXECUTION.parent))
            from execution.download_reports import load_config as dl_load_config
            from execution.download_reports import (
                find_axon_page,
                launch_context,
                maybe_login,
                run_report,
            )

            if mode in ("all", "download"):
                self._log("[1/2] Downloading reports from Axon...")
                try:
                    dl_cfg = dl_load_config(Path(config))
                    downloads_dir = Path(dl_cfg["downloads"]["directory"])
                    downloads_dir.mkdir(parents=True, exist_ok=True)

                    self._log("[INFO] Connecting to browser...")
                    context, we_launched = launch_context(dl_cfg)
                    self._log("[INFO] Browser connected")

                    try:
                        page = find_axon_page(context, dl_cfg["auth"]["base_url"])
                        page.set_viewport_size({"width": 1600, "height": 1000})
                        self._log("[INFO] Logging in to Axon...")
                        maybe_login(page, dl_cfg)
                        self._log("[INFO] Logged in to Axon")

                        self._downloaded_files: list[Path] = []
                        for report in dl_cfg["reports"]:
                            if report.get("enabled", True) is False:
                                continue
                            self._log(f"[INFO] Downloading {report['name']}...")
                            try:
                                result = run_report(page, report, downloads_dir)
                                if result:
                                    self._downloaded_files.append(result)
                                    self._log(f"[OK] Downloaded {report['name']}: {result.name}")
                            except Exception as exc:
                                self._log(f"[ERROR] Failed to download {report['name']}: {exc}")
                                if mode == "download":
                                    return
                    finally:
                        if we_launched:
                            try:
                                context.close()
                            except Exception:
                                pass

                except Exception as exc:
                    self._log(f"[ERROR] Download failed: {exc}")
                    if mode == "download":
                        return
                    build_errors = self._preflight_build_inputs(cfg)
                    if build_errors:
                        for error in build_errors:
                            self._log(f"[ERROR] {error}")
                        self._log("[ERROR] Build step skipped because required inputs are missing.")
                        return
                    self._log("[WARN] Download failed, but required local CSVs already exist. Continuing with build.")

            if mode in ("all", "build"):
                step = "2/2" if mode == "all" else "1/1"
                build_errors = self._preflight_build_inputs(cfg)
                if build_errors:
                    for error in build_errors:
                        self._log(f"[ERROR] {error}")
                    self._log("[ERROR] Build preflight failed.")
                    return
                self._log(f"[{step}] Building Ryan report...")

                dl_dir = cfg.get("downloads", {}).get("directory", "")
                if not dl_dir:
                    dl_dir = str(Path.home() / "Downloads" / "ryan-moves-and-tests")
                historical = cfg.get("historical_ryan", "") or str(Path(dl_dir) / "2026 RYAN MOVES.csv")
                if historical and not Path(historical).exists():
                    result = self._request_missing_file(historical)
                    if result == "cancel":
                        self._log("[INFO] Build cancelled.")
                        return
                    elif result == "new":
                        self._log("[INFO] Starting fresh -- all orders will be treated as new.")
                        Path(historical).parent.mkdir(parents=True, exist_ok=True)
                        Path(historical).write_text("", encoding="utf-8")
                    else:
                        # User picked a new file path
                        historical = result
                        cfg["historical_ryan"] = historical
                        self.save_config(cfg)
                        self._log(f"[INFO] Updated historical file: {Path(historical).name}")
                fresh_output = str(Path(dl_dir) / "generated-ryan-report-latest-new-only.csv")
                append_output = str(Path(dl_dir) / "append-ryan-report-latest.csv")

                from execution.build_ryan_report import main as build_main
                build_args = [
                    "--input-dir", dl_dir,
                    "--output", fresh_output,
                    "--append-to", historical,
                    "--append-output", append_output,
                    "--historical-ryan", historical,
                    "--only-new-orders",
                ]
                try:
                    # build_ryan_report.main() uses argparse -- patch sys.argv
                    old_argv = sys.argv
                    sys.argv = ["build_ryan_report"] + build_args
                    try:
                        build_main()
                    finally:
                        sys.argv = old_argv
                    self._log("[OK] Build complete")
                except SystemExit as exc:
                    if exc.code and exc.code != 0:
                        self._log(f"[ERROR] Build failed (exit {exc.code})")
                        return
                except Exception as exc:
                    self._log(f"[ERROR] Build failed: {exc}")
                    return

                # Append new rows to the xlsx file if it exists.
                self._append_to_xlsx(historical, fresh_output)

            self._log("[DONE] Pipeline complete!")

            # Increment hours saved on success.
            self._increment_hours_saved(3.0)

            # Clean up: delete only the source files downloaded in THIS run.
            downloaded = getattr(self, "_downloaded_files", [])
            if downloaded:
                for f in downloaded:
                    try:
                        if f.exists():
                            f.unlink()
                            self._log(f"  Cleaned up: {f.name}")
                    except OSError:
                        pass

            # List output files.
            dl_dir = Path(cfg.get("downloads", {}).get("directory", ""))
            if dl_dir.exists():
                for f in sorted(dl_dir.iterdir()):
                    if f.suffix == ".csv":
                        size_kb = f.stat().st_size / 1024
                        self._log(f"  {f.name} ({size_kb:.0f} KB)")

            # Push to Airtable if configured.
            airtable = cfg.get("airtable", {})
            if airtable.get("enabled") and airtable.get("token") and airtable.get("table_url"):
                self._push_to_airtable(cfg)

        except Exception as e:
            self._log(f"[ERROR] {e}")
        finally:
            self._running = False

    # -- Airtable --

    def _parse_airtable_url(self, url: str) -> tuple[str, str]:
        """Extract base ID and table ID from an Airtable URL.

        Airtable IDs: base = app + 17 alphanum, table = tbl + 14 alphanum.
        URL format: https://airtable.com/appXXX.../tblYYY.../...
        """
        import re
        # Match app and tbl IDs anywhere in the string.
        base_m = re.search(r"(app[A-Za-z0-9]{14,21})", url)
        table_m = re.search(r"(tbl[A-Za-z0-9]{14,21})", url)
        base_id = base_m.group(1) if base_m else ""
        table_id = table_m.group(1) if table_m else ""
        return base_id, table_id

    def _push_to_airtable(self, cfg: dict) -> None:
        """Push the latest generated report rows to Airtable.

        Uses the Airtable Web API v0 with Personal Access Token auth.
        - Max 10 records per batch (API limit)
        - 300ms delay between batches (rate limit: 5 req/sec/base)
        - Field names are CASE SENSITIVE and must match the table schema
        - 429 responses trigger a 30s retry wait
        """
        import csv as csv_mod
        import urllib.request
        import urllib.error

        airtable = cfg.get("airtable", {})
        token = airtable.get("token", "")
        base_id, table_id = self._parse_airtable_url(airtable.get("table_url", ""))

        if not token:
            self._log("[WARN] No Airtable token configured -- skipping push")
            return
        if not base_id or not table_id:
            self._log("[ERROR] Could not parse base/table ID from Airtable URL. "
                      "URL should look like: https://airtable.com/appXXX/tblYYY")
            return

        # Read the generated report CSV.
        dl_dir = Path(cfg.get("downloads", {}).get("directory", ""))
        report_csv = dl_dir / "generated-ryan-report-latest-new-only.csv"
        if not report_csv.exists():
            self._log("[WARN] No generated report found -- skipping Airtable push")
            return

        # Parse CSV -- skip the two header rows, read data rows.
        with report_csv.open("r", encoding="utf-8-sig") as f:
            lines = list(csv_mod.reader(f))

        if len(lines) < 3:
            self._log("[INFO] No data rows to push to Airtable")
            return

        # CSV column index -> our internal name.
        csv_columns = ["Row", "Truck#", "PO#", "By Whom", "Date Move",
                       "Machine#", "Hour Meter", "Machine Description",
                       "From", "To", "Order#"]

        # Default mapping: our column name -> Airtable field name.
        # Users can override this in config via airtable.field_map.
        default_map = {
            "Truck#": "Truck #",
            "PO#": "PO#",
            "By Whom": "By Whom",
            "Date Move": "Date Move",
            "Machine#": "Machine#",
            "Hour Meter": "Hour Meter",
            "Machine Description": "Machine Description",
            "From": "From Job#",
            "To": "To Job #",
            "Order#": "Order #",
        }
        field_map = airtable.get("field_map", default_map)

        # Which columns to push (configurable, defaults to all mapped ones).
        selected = airtable.get("columns", list(field_map.keys()))

        records = []
        for row in lines[2:]:  # Skip 2 header rows.
            if not row or not any(row):
                continue
            fields = {}
            for i, col_name in enumerate(csv_columns):
                if i < len(row) and col_name in selected and col_name in field_map:
                    airtable_field = field_map[col_name]
                    val = row[i].strip()
                    if val and val.lower() not in ("n/a", "na", "none", "null"):
                        # Date fields need ISO format for Airtable.
                        if airtable_field == "Date Move":
                            try:
                                from datetime import datetime
                                dt = datetime.strptime(val, "%d-%b")
                                dt = dt.replace(year=datetime.now().year)
                                val = dt.strftime("%Y-%m-%d")
                            except ValueError:
                                pass  # Send as-is if parsing fails
                        fields[airtable_field] = str(val)
            if fields:
                records.append({"fields": fields})

        if not records:
            self._log("[INFO] No records to push to Airtable")
            return

        self._log(f"[INFO] Pushing {len(records)} rows to Airtable...")

        # Airtable API: max 10 records per request, 5 requests/sec/base.
        api_url = f"https://api.airtable.com/v0/{base_id}/{table_id}"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        }

        pushed = 0
        for batch_num, i in enumerate(range(0, len(records), 10)):
            batch = records[i:i+10]
            body = json.dumps({"records": batch}).encode()
            req = urllib.request.Request(api_url, data=body, headers=headers, method="POST")

            try:
                with urllib.request.urlopen(req, timeout=30) as resp:
                    pushed += len(batch)
                    self._log(f"  Batch {batch_num+1}: {len(batch)} records sent")
            except urllib.error.HTTPError as e:
                err_body = e.read().decode()[:300]
                if e.code == 429:
                    self._log("[WARN] Rate limited by Airtable -- waiting 30s...")
                    time.sleep(30)
                    # Retry this batch.
                    try:
                        req2 = urllib.request.Request(api_url, data=body, headers=headers, method="POST")
                        with urllib.request.urlopen(req2, timeout=30) as resp2:
                            pushed += len(batch)
                    except Exception as e2:
                        self._log(f"[ERROR] Retry failed: {e2}")
                        return
                elif e.code == 422:
                    # Field name mismatch -- parse the error for the user.
                    self._log(f"[ERROR] Airtable rejected the data (422). This usually means "
                              f"your Airtable table field names don't match exactly.")
                    self._log(f"  Expected fields: {', '.join(selected)}")
                    self._log(f"  Airtable says: {err_body}")
                    self._log(f"  Fix: Make sure your Airtable table has columns with these "
                              f"EXACT names (case-sensitive).")
                    return
                elif e.code == 401:
                    self._log("[ERROR] Airtable token is invalid or expired. "
                              "Create a new one at airtable.com/create/tokens")
                    return
                elif e.code == 403:
                    self._log("[ERROR] Token doesn't have access to this base. "
                              "Edit your token at airtable.com and add this base.")
                    return
                else:
                    self._log(f"[ERROR] Airtable error ({e.code}): {err_body}")
                    return
            except Exception as e:
                self._log(f"[ERROR] Airtable push failed: {e}")
                return

            # Rate limit: wait 300ms between batches.
            if i + 10 < len(records):
                time.sleep(0.3)

        self._log(f"[OK] Pushed {pushed} rows to Airtable")

    # -- Report Management --

    def get_reports(self) -> list[dict]:
        """Return list of configured reports with name and enabled status."""
        cfg = self.load_config()
        reports = cfg.get("reports", [])
        return [{"name": r.get("name", ""), "enabled": r.get("enabled", True)} for r in reports]

    def toggle_report(self, name: str, enabled: bool) -> str:
        """Enable or disable a report by name."""
        cfg = self.load_config()
        for r in cfg.get("reports", []):
            if r.get("name") == name:
                r["enabled"] = enabled
                self.save_config(cfg)
                return "ok"
        return "not found"

    def add_report(self, name: str, menu_path: str) -> str:
        """Add a new report. menu_path is like 'Trucking > Reporter Reports > My Report'."""
        cfg = self.load_config()
        parts = [p.strip() for p in menu_path.split(">")]
        if len(parts) < 2:
            return "Need at least 2 menu items (e.g. 'Trucking > Report Name')"

        steps = [{"action": "click_tab", "name": "Contents", "wait_ms": 1000}]
        # First part is always a tab.
        steps.append({"action": "click_tab", "name": parts[0], "wait_ms": 2000})
        # Middle parts are menu clicks.
        for part in parts[1:-1]:
            steps.append({"action": "click_menu", "text": part, "wait_ms": 3000})
        # Last part is the report itself (also a menu click), then Export.
        steps.append({"action": "click_menu", "text": parts[-1], "wait_ms": 3000})
        steps.append({
            "action": "click_button", "text": "Export",
            "wait_ms": 1000, "triggers_download": True, "timeout_ms": 60000,
        })

        report = {"enabled": True, "name": name, "steps": steps}
        cfg.setdefault("reports", []).append(report)
        self.save_config(cfg)
        return "ok"

    def remove_report(self, name: str) -> str:
        """Remove a report by name."""
        cfg = self.load_config()
        cfg["reports"] = [r for r in cfg.get("reports", []) if r.get("name") != name]
        self.save_config(cfg)
        return "ok"

    # -- Output info --

    def get_output_files(self) -> list[dict]:
        cfg = self.load_config()
        dl_dir = Path(cfg.get("downloads", {}).get("directory", ""))
        files = []
        if dl_dir.exists():
            for f in sorted(dl_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
                if f.suffix == ".csv":
                    st = f.stat()
                    files.append({
                        "name": f.name,
                        "path": str(f),
                        "size_kb": round(st.st_size / 1024, 1),
                        "modified": time.strftime("%Y-%m-%d %H:%M", time.localtime(st.st_mtime)),
                    })
        return files

    def get_unresolved_serials(self) -> list[dict]:
        p = STATE / "unresolved_serials.csv"
        if not p.exists():
            return []
        import csv
        with p.open("r") as f:
            reader = csv.DictReader(f)
            return [dict(row) for row in reader][:50]  # cap at 50

    def open_folder(self, path: str) -> None:
        import subprocess
        if sys.platform == "darwin":
            subprocess.Popen(["open", path])
        elif sys.platform == "win32":
            os.startfile(path)
        else:
            subprocess.Popen(["xdg-open", path])

    # -- Scheduling --

    def _plist_path(self) -> Path:
        return Path.home() / "Library/LaunchAgents/com.andrewnaegele.catom.plist"

    def _task_name(self) -> str:
        return "CatomReportSchedule"

    def get_schedule(self) -> dict[str, Any]:
        """Return the current schedule config, or empty if none."""
        cfg = self.load_config()
        return cfg.get("schedule", {})

    def set_schedule(self, enabled: bool, day: str, hour: int, minute: int) -> str:
        """Set or remove the scheduled run. day: 'daily' or 'monday'-'sunday'."""
        import subprocess

        cfg = self.load_config()
        cfg["schedule"] = {
            "enabled": enabled,
            "day": day,
            "hour": hour,
            "minute": minute,
        }
        self.save_config(cfg)

        if not enabled:
            self._remove_schedule()
            return "Schedule removed"

        if sys.platform == "darwin":
            return self._set_launchd_schedule(day, hour, minute)
        elif sys.platform == "win32":
            return self._set_windows_schedule(day, hour, minute)
        return "Scheduling not supported on this platform"

    def _set_launchd_schedule(self, day: str, hour: int, minute: int) -> str:
        import plistlib
        import subprocess

        config = self.get_config_path()
        # In frozen app, launch the app binary; in dev, use python + script.
        if getattr(sys, "frozen", False):
            program_args = ["/Applications/Catom.app/Contents/MacOS/Catom", "--run-scheduled"]
        else:
            program_args = [sys.executable, str(EXECUTION / "run_pipeline.py"), "--config", config]

        # Build the calendar interval.
        cal: dict[str, int] = {"Hour": hour, "Minute": minute}
        day_map = {
            "sunday": 0, "monday": 1, "tuesday": 2, "wednesday": 3,
            "thursday": 4, "friday": 5, "saturday": 6,
        }
        if day.lower() in day_map:
            cal["Weekday"] = day_map[day.lower()]

        plist = {
            "Label": "com.andrewnaegele.catom",
            "ProgramArguments": program_args,
            "StartCalendarInterval": cal,
            "WorkingDirectory": str(APP_ROOT),
            "StandardOutPath": str(Path.home() / "Library/Logs/catom-report.log"),
            "StandardErrorPath": str(Path.home() / "Library/Logs/catom-report.log"),
            "RunAtLoad": False,
        }

        plist_path = self._plist_path()
        # Unload existing if present.
        subprocess.run(["launchctl", "unload", str(plist_path)],
                       capture_output=True, check=False)

        with plist_path.open("wb") as f:
            plistlib.dump(plist, f)

        subprocess.run(["launchctl", "load", str(plist_path)], check=True)
        return f"Scheduled: {day} at {hour:02d}:{minute:02d}"

    def _set_windows_schedule(self, day: str, hour: int, minute: int) -> str:
        import subprocess

        config = self.get_config_path()
        if getattr(sys, "frozen", False):
            program = f'"{sys.executable}" "--run-scheduled"'
        else:
            program = f'"{sys.executable}" "{EXECUTION / "run_pipeline.py"}" "--config" "{config}"'
        task = self._task_name()
        time_str = f"{hour:02d}:{minute:02d}"

        # Delete existing task if present.
        subprocess.run(["schtasks", "/delete", "/tn", task, "/f"],
                       capture_output=True, check=False)

        if day.lower() == "daily":
            schedule_type, day_arg = "/sc", "daily"
        else:
            schedule_type = "/sc"
            day_arg = "weekly"

        cmd = [
            "schtasks", "/create", "/tn", task,
            "/tr", program,
            "/sc", day_arg,
            "/st", time_str,
        ]
        if day.lower() != "daily":
            cmd.extend(["/d", day[:3].upper()])

        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            return f"Failed to create schedule: {result.stderr}"
        return f"Scheduled: {day} at {time_str}"

    def _remove_schedule(self) -> None:
        import subprocess
        if sys.platform == "darwin":
            plist_path = self._plist_path()
            subprocess.run(["launchctl", "unload", str(plist_path)],
                           capture_output=True, check=False)
            plist_path.unlink(missing_ok=True)
        elif sys.platform == "win32":
            subprocess.run(["schtasks", "/delete", "/tn", self._task_name(), "/f"],
                           capture_output=True, check=False)

    # -- AI Troubleshooting --

    def troubleshoot(self, error_log: str) -> str:
        """Send error logs to Claude API for troubleshooting advice."""
        try:
            import urllib.request
            import urllib.error

            # Look for API key in config or environment.
            cfg = self.load_config()
            api_key = cfg.get("anthropic_api_key", "") or os.environ.get("ANTHROPIC_API_KEY", "")
            if not api_key:
                return (
                    "No API key configured. To enable AI troubleshooting:\n\n"
                    "1. Go to Settings\n"
                    "2. Add your Anthropic API key\n"
                    "3. Try again\n\n"
                    "Common fixes:\n"
                    "- 'Profile lock' error: Close your browser, reopen it, try again\n"
                    "- 'element not found' error: The Axon page layout may have changed -- contact support\n"
                    "- 'Download failed' error: Make sure you're logged into Axon in your browser\n"
                    "- 'FileNotFoundError': Check that your Ryan Moves CSV path is correct in Settings"
                )

            prompt = (
                "You are a troubleshooting assistant for the Ryan Report app. "
                "This app downloads reports from Axon TMS (a trucking management system) "
                "via browser automation (Playwright CDP) and builds a combined CSV report. "
                "The user got an error. Diagnose the problem and give a clear, "
                "non-technical fix in 2-3 sentences. Here's the error log:\n\n"
                f"{error_log[-2000:]}"
            )

            body = json.dumps({
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 300,
                "messages": [{"role": "user", "content": prompt}],
            }).encode()

            req = urllib.request.Request(
                "https://api.anthropic.com/v1/messages",
                data=body,
                headers={
                    "Content-Type": "application/json",
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                },
            )

            with urllib.request.urlopen(req, timeout=15) as resp:
                result = json.loads(resp.read())
                return result["content"][0]["text"]

        except urllib.error.HTTPError as e:
            return f"API error ({e.code}): Check that your API key is valid."
        except Exception as e:
            return f"Could not reach Claude: {e}\n\nCheck your internet connection and try again."


# ---------------------------------------------------------------------------
# App entry point.
# ---------------------------------------------------------------------------

def main() -> None:
    # --run-scheduled: headless mode for cron/launchd/Task Scheduler.
    if "--run-scheduled" in sys.argv:
        api = PipelineAPI()
        api._run("all")
        sys.exit(0)

    api = PipelineAPI()
    window = webview.create_window(
        "Catom",
        _ui_path(),
        js_api=api,
        width=800,
        height=620,
        resizable=True,
        min_size=(600, 400),
    )
    api.set_window(window)
    webview.start(debug=("--debug" in sys.argv))


if __name__ == "__main__":
    main()
