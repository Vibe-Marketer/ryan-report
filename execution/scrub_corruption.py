"""One-shot scrub: clear paste-typo corruption from page-footer rows.

Some old page-footer rows in the canonical workbooks have stray paste-typo
strings in cols B-P (e.g. col I has `35equipment@ryancentral.com0.2-NEW
CARLISLE`). A page-footer row should have ONLY:
  - col A: the equipment email
  - col H: the "PAGE: NN OF 2026" string

Everything else should be empty. This script:
  1. Walks every sheet in the workbook
  2. For each row whose col A is `equipment@ryancentral.com`, scrubs cols B-P
     EXCEPT col H (which carries the page text)
  3. Saves the workbook only if changes were made

Usage:
    python -m execution.scrub_corruption /path/to/2026\\ RYAN\\ MOVES.xlsx [more.xlsx ...]
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

EMAIL_TEXT = "equipment@ryancentral.com"
PAGE_COL = 8  # col H


def scrub(path: Path) -> int:
    """Scrub one workbook in place. Returns number of cells cleared."""
    wb = load_workbook(path)
    cleared = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            if not row:
                continue
            cell_a = row[0]
            v = cell_a.value
            if not isinstance(v, str) or EMAIL_TEXT not in v:
                continue
            # Found a page-footer row.
            # Normalize col A back to the canonical email (in case it had
            # extra junk concatenated like '35equipment@ryancentral.com').
            if v != EMAIL_TEXT:
                cell_a.value = EMAIL_TEXT
                cleared += 1
            # Clear cols B-P except H (page text). Clear value AND hyperlink
            # — older corruption had the typo'd email both as the cell's
            # display value and as a mailto: hyperlink target; clearing only
            # one leaves the other behind in Excel's display.
            for cell in row[1:16]:
                if cell.column == PAGE_COL:
                    continue
                if cell.value not in (None, ""):
                    cell.value = None
                    cleared += 1
                if cell.hyperlink is not None:
                    cell.hyperlink = None
                    cleared += 1
            # Col A may also have a malformed mailto: target
            try:
                hl = cell_a.hyperlink
                if hl is not None and getattr(hl, "target", None) not in (
                    None, "", f"mailto:{EMAIL_TEXT}"
                ):
                    hl.target = f"mailto:{EMAIL_TEXT}"
                    cleared += 1
            except Exception:
                pass
    if cleared:
        wb.save(path)
    return cleared


def main() -> None:
    if len(sys.argv) < 2:
        print("usage: python -m execution.scrub_corruption FILE.xlsx [...]")
        sys.exit(2)
    total = 0
    for arg in sys.argv[1:]:
        p = Path(arg)
        if not p.exists():
            print(f"[SKIP] not found: {p}")
            continue
        n = scrub(p)
        print(f"[OK] {p}  cleared {n} cell(s)")
        total += n
    print(f"--- total: {total} cells cleared across {len(sys.argv) - 1} file(s)")


if __name__ == "__main__":
    main()
