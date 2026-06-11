from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

TARGET_HEADER_ROW_1 = [
    "",
    "Truck",
    "",
    "By",
    "Date",
    "",
    "Hour",
    "Machine",
    "From",
    "To",
    "Order #",
    "",
    "",
    "",
    "",
    "",
]
TARGET_HEADER_ROW_2 = [
    "",
    "#",
    "PO#",
    "Whom",
    "Move",
    "Machine#",
    "Meter",
    "Description",
    "Job#  ",
    "Job#                          ",
    "",
    "",
    "",
    "",
    "",
    "",
]
SERIAL_COLUMNS = ["Serial #", "Serial #2", "Serial #3", "Serial #4"]
IGNORED_SERIAL_VALUES = {"", "0", "0.0", "00", "o", "na", "n/a", "none", "null"}
ORDER_MASTER_NAME = "Order Master Report"
NEW_RYAN_NAME = "New RYAN"


@dataclass
class OrderMasterRecord:
    order_number: str
    move_date: str
    origin: str
    destination: str
    driver_name: str


@dataclass
class GeneratedRow:
    po: str
    driver_initials: str  # Truck column (e.g. 'RH' from driver 'Hendrickson, Rod')
    move_date: str
    machine: str
    meter: str
    description: str
    origin: str
    destination: str
    order_number: str
    by_whom: str = "DR"  # By Whom column — initials of the Ryan dispatcher who called the order in


def read_csv_rows(path: Path) -> list[list[str]]:
    # Handle xlsx files natively. Reads only the active sheet — for
    # multi-sheet historical workbooks, prefer read_historical_rows below.
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([_xlsx_cell_to_str(cell) for cell in row])
            wb.close()
            return rows
        except ImportError:
            raise RuntimeError(
                f"Cannot read {path.name}: openpyxl is required for xlsx files. "
                f"Install it with: pip install openpyxl"
            )
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.reader(handle))
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("csv", b"", 0, 1, f"Unable to decode CSV file: {path}")


def _xlsx_cell_to_str(cell) -> str:
    """Render an openpyxl cell value as a CSV-equivalent string.

    Datetimes are written as the same '%-d-%b' format the CSV pipeline emits
    (e.g. '6-Jan'), so dedup keys built off these strings line up with rows
    produced from CSV inputs.
    """
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        try:
            return cell.strftime("%-d-%b")
        except ValueError:
            # Windows: %-d not supported; %#d is the equivalent.
            return cell.strftime("%#d-%b")
    return str(cell)


def read_historical_rows(path: Path) -> list[list[str]]:
    """Load rows for the historical equipment lookup.

    For an .xlsx workbook, unions rows from every sheet whose name is a 4-digit
    year (e.g. '2016' through '2026'), so the serial -> description / meter
    lookup sees the full equipment history that drives `Ryan` description
    fill-in. Falls back to the single-sheet read_csv_rows for CSV inputs and
    for xlsx workbooks that have no year-named sheets.
    """
    if path.suffix.lower() not in (".xlsx", ".xls"):
        return read_csv_rows(path)

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError(
            f"Cannot read {path.name}: openpyxl is required for xlsx files. "
            f"Install it with: pip install openpyxl"
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    year_sheets = [name for name in wb.sheetnames if name.isdigit() and len(name) == 4]
    if not year_sheets:
        wb.close()
        return read_csv_rows(path)

    rows: list[list[str]] = []
    for name in year_sheets:
        ws = wb[name]
        for row in ws.iter_rows(values_only=True):
            rows.append([_xlsx_cell_to_str(cell) for cell in row])
    wb.close()
    return rows


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


# Standard Catom serial: NN-NNNNN or NN NNNNN (2-3 digit prefix, 3-6 digit body).
_SERIAL_RE = re.compile(r'\b(\d{2,3})[-\s](\d{3,6})\b')
# VIN-like single token (alphanumeric, >= 7 chars, contains a digit, no spaces).
_VIN_RE = re.compile(r'^[A-Za-z0-9][A-Za-z0-9\-]{6,}$')


def normalize_serial(value: str) -> str:
    """Extract a clean equipment serial from the Order Master cell.

    Handles the messy real-world cases Eric flagged where attachment serials
    arrive wrapped in words or with spaces instead of dashes:

      '87-16453 BUCKET'   -> '87-16453'
      'bucket 87-16267'   -> '87-16267'
      'forks 87-11569'    -> '87-11569'
      '52 12711'          -> '52-12711'
      'RIC 99-12034'      -> '99-12034'

    Pure descriptive junk ('12 barricades', 'hrs. 4593.2', 'p', '6') and the
    ignored placeholders ('0', 'N/A') return '' so they never reach the report.
    VIN-style serials ('1FF035GXCNK30', 'GS300091') are kept as-is.
    """
    value = normalize_space(value)
    if not value or value.lower() in IGNORED_SERIAL_VALUES:
        return ""
    # 1) Standard Catom serial embedded anywhere -> normalize to NN-NNNNN.
    m = _SERIAL_RE.search(value)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    # 2) VIN-like single token (no spaces, has a digit) -> keep, uppercased.
    if " " not in value and _VIN_RE.match(value) and any(c.isdigit() for c in value):
        return value.upper()
    # 3) Everything else is descriptive text, not a serial.
    return ""


def normalize_meter(value: str, default: str = "N/A") -> str:
    value = normalize_space(value)
    if not value or value.lower() in IGNORED_SERIAL_VALUES:
        return default
    return value


def format_move_date(value: str) -> str:
    value = normalize_space(value)
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).strftime("%-d-%b")
        except ValueError:
            continue
    return value


def derive_driver_initials(driver_name: str) -> str:
    driver_name = normalize_space(driver_name)
    if not driver_name:
        return ""
    if "," in driver_name:
        last, first = [normalize_space(part) for part in driver_name.split(",", 1)]
        return (first[:1] + last[:1]).upper()
    parts = driver_name.split(" ")
    if len(parts) >= 2:
        return (parts[0][:1] + parts[-1][:1]).upper()
    return driver_name[:2].upper()


def discover_latest_file(input_dir: Path, prefix: str) -> Path:
    matches = sorted(
        input_dir.glob(f"{prefix}*.csv"), key=lambda p: p.stat().st_mtime, reverse=True
    )
    if not matches:
        raise FileNotFoundError(
            f"No file matching '{prefix}*.csv' found in {input_dir}"
        )
    return matches[0]


def parse_order_master(path: Path) -> dict[str, OrderMasterRecord]:
    rows = read_csv_rows(path)
    start_index = (
        next(i for i, row in enumerate(rows) if row and row[0] == "Order#") + 1
    )
    records = {}
    current = None
    for row in rows[start_index:]:
        padded = row + [""] * (10 - len(row))
        first = normalize_space(padded[0])
        location = normalize_space(padded[2])
        if first.isdigit():
            current = OrderMasterRecord(
                first,
                format_move_date(padded[1]),
                location,
                "",
                normalize_space(padded[4]),
            )
            records[first] = current
            continue
        if not current:
            continue
        if first:
            current = None
            continue
        if (
            not current.destination
            and location
            and location
            not in {
                "Freight Amount",
                "Other Charges",
                "Total Billed",
                "TotalBilled",
                "Total Payout",
                "Net Revenue",
            }
        ):
            current.destination = location
    return records


def normalize_location(value: str) -> str:
    """Stopgap From/To formatter.

    Customer wants `<job#>-<CITY>` (e.g. `2503.2-HOFFMAN ESTATES`). Until the
    Distribution PDF parser ships the asset->job# crosswalk, strip the trailing
    state code (`, IL` / `, WI`) and uppercase. So Axon's `"HOFFMAN ESTATES, IL"`
    becomes `"HOFFMAN ESTATES"` instead of being shipped to the customer with
    state codes that no other row has. The job# prefix gets added back when
    the Distribution PDF lookup lands.
    """
    s = normalize_space(value or "")
    if not s:
        return ""
    # Strip trailing ", ST" (2-letter state) — match real Axon output.
    s = re.sub(r",\s*[A-Z]{2}\s*$", "", s, flags=re.IGNORECASE)
    return s.upper()


def parse_distribution_pdf(pdf_path: Path) -> dict[str, str]:
    """Parse Ryan's RIC EM Distribution Report PDF -> {asset: '<job>-<CITY>'}.

    Each Job/Location section header gives the job# + city. Asset rows under
    that section ('NN-NNNNN  description  vin...') get mapped to that section
    until the next header. The consolidated index pages at the end of the PDF
    use a different layout and are skipped — front-of-doc regional listings
    are sufficient.

    Returns a flat dict where the value is already in '<job>-<CITY>' or shop
    label form, ready to drop into From/To.
    """
    import re as _re

    if not pdf_path.exists():
        return {}

    # Pure-Python extraction via pdfplumber. Earlier versions shelled out
    # to `pdftotext`, but that binary isn't on Windows by default and our
    # PyInstaller bundle doesn't ship it — so the extraction silently
    # failed on Eric's machine. pdfplumber ships in the bundle and produces
    # identical asset/job mappings against the same PDF (validated 2,826
    # assets matched 1:1 between pdftotext and pdfplumber).
    try:
        import pdfplumber  # type: ignore
    except ImportError:
        # No PDF library available; degrade gracefully.
        return {}

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            text = "\n".join((page.extract_text(layout=True) or "") for page in pdf.pages)
    except Exception:
        return {}
    lines = text.splitlines()

    # Detect where the consolidated 2-column index begins; only parse before it.
    index_line = _re.compile(
        r'^\s*\d+\s*-\s*\d+\s+\S.*\s\d{3,5}(?:\.\d)?-\s+\d+\s*-\s*\d+\s+'
    )
    index_start = next(
        (i for i, l in enumerate(lines) if index_line.match(l)),
        len(lines),
    )

    job_hdr = _re.compile(
        r'^Job:\s+(\d+(?:\.\d+)?)-\s*(.+?),\s*([A-Z]{2})(?:\b|\s)'
    )
    loc_hdr = _re.compile(r'^Location:\s+(.+)')
    cont_re = _re.compile(r'\s*-\s*Continued\s*$')
    asset_re = _re.compile(r'^(\d{1,3})\s*-\s*(\d{3,5})\b')

    result: dict[str, str] = {}
    current: str | None = None
    for ln in range(index_start):
        line = lines[ln].strip()
        if not line:
            continue
        m = job_hdr.match(line)
        if m:
            body = cont_re.sub("", m.group(2)).strip()
            city = body.rsplit(' - ', 1)[-1] if ' - ' in body else body
            current = f"{m.group(1)}-{city.strip().upper()}"
            continue
        m = loc_hdr.match(line)
        if m:
            head = cont_re.sub("", m.group(1)).split(',')[0].strip()
            current = _re.sub(r'\s*-\s*', '-', head).upper()
            continue
        if current is None:
            continue
        m_a = asset_re.match(line)
        if m_a:
            asset = f"{m_a.group(1)}-{m_a.group(2)}"
            # First mapping wins (regional pages); index pages were skipped.
            if asset not in result:
                result[asset] = current
    return result


def discover_distribution_pdf(input_dir: Path) -> Path | None:
    """Find the newest RIC EM Distribution Report PDF in the source folder.

    Eric drops the PDF into the SAME folder as the Axon CSVs (no separate
    folder). We pick the newest by mtime; older PDFs are kept for diff-based
    From-side resolution in a future phase.
    """
    if not input_dir.is_dir():
        return None
    candidates = sorted(
        list(input_dir.glob("Equipment Distribution Report*.pdf"))
        + list(input_dir.glob("RIC EM Distribution*.pdf")),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def parse_historical_orders(path: Path) -> dict[str, tuple[str, str]]:
    """Build Order# -> (from_str, to_str) crosswalk from the historical xlsx.

    Reads every year-sheet and records the From/To values seen for each
    Order#. Used as the primary lookup for From/To formatting — when a
    repeat order shows up in Axon, we use the exact From/To string Eric
    has used historically (which already has the `<job#>-<CITY>` format).
    Falls back to normalize_location() for orders not seen historically.
    """
    crosswalk: dict[str, tuple[str, str]] = {}
    rows = read_historical_rows(path)
    for row in rows:
        padded = row + [""] * 11
        order = normalize_space(padded[10])
        if not order.isdigit():
            continue
        from_v = normalize_space(padded[8])
        to_v = normalize_space(padded[9])
        if from_v or to_v:
            crosswalk[order] = (from_v, to_v)
    return crosswalk


def parse_historical_ryan(path: Path) -> tuple[dict[str, dict[str, str]], Counter[str]]:
    rows = read_historical_rows(path)
    description_counts = defaultdict(Counter)
    meter_counts = defaultdict(Counter)
    for row in rows:
        padded = row + [""] * 11
        if not normalize_space(padded[0]).isdigit() or not normalize_space(padded[10]):
            continue
        serial = normalize_serial(padded[5])
        if not serial:
            continue
        desc = normalize_space(padded[7])
        meter = normalize_meter(padded[6])
        if desc:
            description_counts[serial][desc] += 1
        if meter:
            meter_counts[serial][meter] += 1
    lookup = {}
    for serial, counts in description_counts.items():
        lookup[serial] = {
            "description": counts.most_common(1)[0][0],
            "meter": meter_counts[serial].most_common(1)[0][0]
            if meter_counts[serial]
            else "N/A",
        }
    return lookup, Counter({s: sum(c.values()) for s, c in description_counts.items()})


def derive_called_initials(called: str) -> str:
    """Extract initials from Summary's 'Called' field.

    The field looks like 'Doug Reibel 815-405-8273' — strip the phone trailing
    portion, then return first-letter-of-first-name + first-letter-of-last-name.
    'Doug Reibel ...' -> 'DR'. Falls back to 'DR' (Eric's most common dispatcher,
    matches historical pattern) if parsing fails.
    """
    s = normalize_space(called)
    if not s:
        return ""
    # Strip a trailing phone number like '815-405-8273' or '(815) 405-8273'
    s = re.sub(r'\s*[\(\d][\d\-\)\.\s]{6,}\d\s*$', '', s).strip()
    if not s:
        return ""
    parts = s.split()
    if len(parts) >= 2:
        return (parts[0][:1] + parts[-1][:1]).upper()
    return parts[0][:2].upper() if parts else ""


def format_from_to(text: str, city: str, customer_job_fallback: str = "") -> str:
    """Build the Ryan-format From/To column value from a Shipper or Consignee
    free-text field plus its City field.

    Single rule:
      - If the text starts with `<digits>(.<digits>)?` -> '<digits>-<UPPER CITY>'
      - Else if customer_job_fallback starts with digits and Consignee text didn't,
        use that prefix.
      - Else (shop name like 'RYAN SHOP', 'MACHOLLS', 'ALTORFER MOKENA') ->
        '<UPPER text>-<UPPER city>'.
    """
    text = normalize_space(text)
    city = normalize_space(city).upper()
    if not text and not city:
        # Last-resort fallback: use the customer_job alone if present.
        cj = normalize_space(customer_job_fallback)
        return cj.upper()
    if not text:
        cj = normalize_space(customer_job_fallback)
        if cj:
            m = re.match(r'^(\d+(?:\.\d+)?)', cj)
            if m and city:
                return f"{m.group(1)}-{city}"
        return city
    # Try to extract a leading job# (e.g. '2560.2', '3416.2', '1919', '31100').
    m = re.match(r'^(\d+(?:\.\d+)?)', text)
    if m:
        return f"{m.group(1)}-{city}" if city else m.group(1)
    # No job# in the text. Try the customer_job fallback.
    cj = normalize_space(customer_job_fallback)
    cj_match = re.match(r'^(\d+(?:\.\d+)?)', cj)
    if cj_match:
        return f"{cj_match.group(1)}-{city}" if city else cj_match.group(1)
    # Shop / yard name. Just upper-case the text + city.
    return f"{text.upper()}-{city}" if city else text.upper()


def classify_order_master_csv(path: Path) -> str:
    """Sniff an Order Master Report CSV. Returns 'summary', 'detail', or 'unknown'.

    Both presets export with the same filename prefix; we tell them apart by
    looking at the layout marker row near the top:
      - Detail preset:  contains a row with 'Detail' as the only non-empty cell
      - Summary preset: contains a row with 'Summary Per Order'
    Or by the header row that follows: Detail starts with 'Order#', Summary
    starts with 'PO#'.
    """
    try:
        rows = read_csv_rows(path)
    except Exception:
        return "unknown"
    for row in rows[:12]:
        joined = "|".join(normalize_space(c) for c in row).lower()
        if "summary per order" in joined or "po#|called|" in joined:
            return "summary"
        if "|detail|" in f"|{joined}|" or joined.startswith("order#|end date|"):
            return "detail"
    return "unknown"


def discover_order_master_pair(input_dir: Path) -> tuple[Path | None, Path | None]:
    """Find the newest Detail + Summary CSV pair in input_dir.

    Returns (detail_path, summary_path). Either can be None if not present.
    """
    if not input_dir.is_dir():
        return None, None
    candidates = sorted(
        input_dir.glob(f"{ORDER_MASTER_NAME}*.csv"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    detail: Path | None = None
    summary: Path | None = None
    for c in candidates:
        kind = classify_order_master_csv(c)
        if kind == "detail" and detail is None:
            detail = c
        elif kind == "summary" and summary is None:
            summary = c
        if detail and summary:
            break
    return detail, summary


def parse_order_master_detail_drivers(path: Path) -> dict[str, str]:
    """Read the Detail report (115229-style) and return {Order#: driver_name}.

    The Detail layout's column 4 'Equipment/Employee' carries the driver name
    (e.g. 'Hendrickson, Rod') on the order's primary row. We only need the
    driver -> initials translation for the Truck column of the output xlsx;
    everything else now comes from the Summary report.
    """
    rows = read_csv_rows(path)
    drivers: dict[str, str] = {}
    # Find the column-header row.
    start_index = None
    for i, row in enumerate(rows):
        if row and normalize_space(row[0]) == "Order#":
            start_index = i + 1
            break
    if start_index is None:
        return drivers
    for row in rows[start_index:]:
        padded = row + [""] * (10 - len(row))
        first = normalize_space(padded[0])
        if first.isdigit():
            employee = normalize_space(padded[4])
            # Skip non-driver entries (some rows show 'Total Billed' etc).
            if employee and "," in employee:
                drivers[first] = employee
    return drivers


def parse_order_master_summary(
    path: Path,
    drivers: dict[str, str] | None = None,
    serial_lookup: dict[str, dict[str, str]] | None = None,
    distribution_lookup: dict[str, str] | None = None,
) -> tuple[list[GeneratedRow], list[dict[str, str]]]:
    """Parse the Order Master Summary report (115319-style) into GeneratedRow.

    Header layout:
        PO#, Called, Order Start Date, Serial #, Serial #2, Serial #3, Serial #4,
        Hour Meter, Shipper, Shipper City, Consignee, Consignee City,
        Customer Job #, Order#, Ready To Invoice

    For each order: serial #1 produces a row with Hour Meter populated; serials
    #2/3/4 (when present) produce additional rows with the same Order#, From,
    To, etc. but a blank Hour Meter (per Eric's pattern: bucket / attachment
    rows have no meter).
    """
    drivers = drivers or {}
    serial_lookup = serial_lookup or {}
    distribution_lookup = distribution_lookup or {}

    rows = read_csv_rows(path)
    # Locate the column-header row (starts with 'PO#').
    start_index = None
    for i, row in enumerate(rows):
        if row and normalize_space(row[0]) == "PO#":
            start_index = i + 1
            break
    if start_index is None:
        return [], []

    generated: list[GeneratedRow] = []
    unresolved: list[dict[str, str]] = []
    current_date = ""

    for row in rows[start_index:]:
        padded = row + [""] * (15 - len(row))
        # Single-cell rows in the Summary export are date sub-headers like
        # '04/27/2026' that group orders under that move date.
        first = normalize_space(padded[0])
        # If this is a single-column date header line, latch it.
        non_empty = [normalize_space(v) for v in padded if normalize_space(v)]
        if len(non_empty) == 1 and re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', non_empty[0]):
            current_date = non_empty[0]
            continue

        order_number = normalize_space(padded[13])
        if not order_number.isdigit():
            continue

        po = normalize_space(padded[0])
        called = normalize_space(padded[1])
        date_field = normalize_space(padded[2]) or current_date
        serials = [
            normalize_serial(padded[3]),
            normalize_serial(padded[4]),
            normalize_serial(padded[5]),
            normalize_serial(padded[6]),
        ]
        hour_meter_raw = normalize_space(padded[7])
        shipper = normalize_space(padded[8])
        shipper_city = normalize_space(padded[9])
        consignee = normalize_space(padded[10])
        consignee_city = normalize_space(padded[11])
        customer_job = normalize_space(padded[12])

        # Format the From/To once per order.
        from_str = format_from_to(shipper, shipper_city)
        to_str = format_from_to(consignee, consignee_city, customer_job_fallback=customer_job)

        # Date: the Summary writes '04/27/2026'; format_move_date converts to '27-Apr'.
        move_date = format_move_date(date_field) if date_field else ""

        # Hour meter — first serial only. '4 hours', '1620', '2411', '0', 'N/A'.
        first_meter = normalize_meter(hour_meter_raw)

        # Driver initials from the Detail report cross-reference.
        driver_initials = derive_driver_initials(drivers.get(order_number, ""))

        # By Whom column comes from Summary's 'Called' field.
        by_whom = derive_called_initials(called)

        # PO column stays blank per Eric's spec — Ryan fills it in later.
        # (We still expose the Summary's PO# as a separate field if a future
        # caller wants it; we just don't put it in the output.)

        for idx, serial in enumerate(serials):
            if not serial:
                continue
            # Description: historical xlsx primary; PDF fallback for new assets.
            desc_entry = serial_lookup.get(serial, {})
            description = desc_entry.get("description", "")
            if not description and serial in distribution_lookup:
                # PDF fallback — distribution_lookup may map to a job-city OR
                # description string depending on future schema. Today it's
                # job-city; we leave description blank if we don't have a real
                # description (better than mislabeling).
                pass
            if not description:
                # Track it so we can backfill the description later...
                unresolved.append({
                    "order_number": order_number,
                    "serial": serial,
                    "source_column": SERIAL_COLUMNS[idx] if idx < len(SERIAL_COLUMNS) else f"Serial#{idx+1}",
                    "issue": "Missing description",
                    "suggested_description": "",
                })
                # ...but DO NOT drop it. Eric's #1 bug: attachment serials with
                # no historical description were silently skipped, so multi-
                # attachment moves only showed the first (described) serial.
                # Emit the serial with a blank description — the machine number
                # is the data that matters; Ryan/Eric fills the description.

            generated.append(GeneratedRow(
                po="",  # blank per spec — Ryan fills this in after receiving the report
                driver_initials=driver_initials,  # Truck column (e.g. 'RH')
                move_date=move_date,
                machine=serial,
                meter=(first_meter if idx == 0 else "N/A"),
                description=description,
                origin=from_str,
                destination=to_str,
                order_number=order_number,
                by_whom=by_whom or "DR",
            ))

    return generated, unresolved


def parse_overrides(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    rows = read_csv_rows(path)
    if not rows:
        return {}
    header = rows[0]
    overrides = {}
    for raw in rows[1:]:
        row = dict(zip(header, raw + [""] * (len(header) - len(raw))))
        serial = normalize_serial(row.get("serial", ""))
        if serial:
            overrides[serial] = {
                "description": normalize_space(row.get("description", "")),
                "meter": normalize_meter(row.get("meter", "N/A")),
            }
    return overrides


def build_serial_lookup(
    generated_seed: dict[str, dict[str, str]],
    historical: dict[str, dict[str, str]],
    overrides: dict[str, dict[str, str]],
) -> dict[str, dict[str, str]]:
    lookup = {**generated_seed, **historical}
    lookup.update(overrides)
    return lookup


def parse_new_ryan(path: Path) -> list[dict[str, str]]:
    rows = read_csv_rows(path)
    if not rows:
        return []
    header = rows[0]
    return [
        dict(zip(header, row + [""] * (len(header) - len(row)))) for row in rows[1:]
    ]


def resolve_description(
    serial: str,
    serial_index: int,
    row: dict[str, str],
    lookup: dict[str, dict[str, str]],
) -> str:
    commodity = normalize_space(row.get("Order Commodity", ""))
    if serial_index == 0 and commodity:
        return commodity
    return lookup.get(serial, {}).get("description", "")


def resolve_meter(
    serial: str,
    serial_index: int,
    row: dict[str, str],
    lookup: dict[str, dict[str, str]],
) -> str:
    if serial_index == 0:
        return normalize_meter(row.get("Hour Meter", ""))
    return lookup.get(serial, {}).get("meter", "N/A") or "N/A"


def collect_generated_rows(
    new_rows: list[dict[str, str]],
    order_master: dict[str, OrderMasterRecord],
    serial_lookup: dict[str, dict[str, str]],
    order_crosswalk: dict[str, tuple[str, str]] | None = None,
    distribution_lookup: dict[str, str] | None = None,
) -> tuple[list[GeneratedRow], list[dict[str, str]]]:
    generated = []
    unresolved = []
    crosswalk = order_crosswalk or {}
    pdf_lookup = distribution_lookup or {}
    for row in new_rows:
        if normalize_space(row.get("Bill To Name", "")) != "Ryan, Inc.":
            continue
        order_number = normalize_space(row.get("Order#", ""))
        if not order_number:
            continue
        master = order_master.get(order_number)
        if not master:
            continue

        # From/To resolution chain:
        #   1. Historical crosswalk (repeat orders -> exact <job>-<CITY> Eric used)
        #   2. Distribution-PDF lookup by asset (asset's current job assignment)
        #   3. Stopgap: state-stripped uppercase city
        # PDF lookup is asset-specific, so it happens inside the SERIAL loop.
        # Order-level crosswalk takes precedence — it's been hand-validated.
        cw = crosswalk.get(order_number)
        if cw and (cw[0] or cw[1]):
            order_from, order_to = cw
            order_resolved = True
        else:
            order_from = normalize_location(master.origin)
            order_to = normalize_location(master.destination)
            order_resolved = False

        for index, column in enumerate(SERIAL_COLUMNS):
            serial = normalize_serial(row.get(column, ""))
            if not serial:
                continue
            description = resolve_description(serial, index, row, serial_lookup)
            meter = resolve_meter(serial, index, row, serial_lookup)
            if not description:
                unresolved.append(
                    {
                        "order_number": order_number,
                        "serial": serial,
                        "source_column": column,
                        "issue": "Missing description",
                        "suggested_description": "",
                    }
                )
                continue

            # PDF lookup — the asset's job# assignment per the latest
            # Distribution Report. Used only when crosswalk didn't resolve
            # (so we don't override Eric's hand-validated history). For new
            # orders, this gives us the <job>-<CITY> format for free.
            if order_resolved:
                from_str, to_str = order_from, order_to
            else:
                pdf_value = pdf_lookup.get(serial)
                if pdf_value:
                    # Asset is at job X per PDF. For now, use the same
                    # PDF value for both From and To (single-PDF case).
                    # Multi-PDF diff for distinct From vs To is a follow-up.
                    from_str = pdf_value
                    to_str = pdf_value
                else:
                    from_str = order_from
                    to_str = order_to

            generated.append(
                GeneratedRow(
                    normalize_space(row.get("Customer PO#", "")),
                    derive_driver_initials(master.driver_name),
                    master.move_date,
                    serial,
                    meter,
                    description,
                    from_str,
                    to_str,
                    order_number,
                )
            )
    return generated, unresolved


def generated_row_key(row: GeneratedRow) -> tuple[str, str, str, str, str]:
    return (row.order_number, row.machine, row.move_date, row.origin, row.destination)


def parse_existing_target_rows(
    path: Path,
) -> tuple[list[list[str]], set[tuple[str, str, str, str, str]], set[str], int]:
    rows = read_csv_rows(path)
    preserved = []
    seen = set()
    seen_orders = set()
    max_line = 0
    for row in rows:
        preserved.append(row)
        padded = row + [""] * 11
        if normalize_space(padded[0]).isdigit():
            max_line = max(max_line, int(normalize_space(padded[0])))
            seen_orders.add(normalize_space(padded[10]))
            seen.add(
                (
                    normalize_space(padded[10]),
                    normalize_serial(padded[5]),
                    normalize_space(padded[4]),
                    normalize_space(padded[8]),
                    normalize_space(padded[9]),
                )
            )
    return preserved, seen, seen_orders, max_line


def filter_new_orders(
    generated_rows: list[GeneratedRow], append_to: Path
) -> tuple[list[GeneratedRow], int]:
    _, _, seen_orders, _ = parse_existing_target_rows(append_to)
    filtered = [row for row in generated_rows if row.order_number not in seen_orders]
    return filtered, len(generated_rows) - len(filtered)


def filter_unresolved_new_orders(
    unresolved: list[dict[str, str]], append_to: Path
) -> list[dict[str, str]]:
    _, _, seen_orders, _ = parse_existing_target_rows(append_to)
    return [row for row in unresolved if row.get("order_number", "") not in seen_orders]


def write_target_csv(path: Path, generated_rows: Iterable[GeneratedRow]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(TARGET_HEADER_ROW_1)
        writer.writerow(TARGET_HEADER_ROW_2)
        for idx, row in enumerate(generated_rows, start=1):
            writer.writerow(
                [
                    idx,
                    row.driver_initials,
                    row.po,
                    row.by_whom or "DR",
                    row.move_date,
                    row.machine,
                    row.meter,
                    row.description,
                    row.origin,
                    row.destination,
                    row.order_number,
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )


def write_append_csv(
    path: Path, append_to: Path, generated_rows: list[GeneratedRow]
) -> tuple[int, int]:
    preserved, seen, _, max_line = parse_existing_target_rows(append_to)
    skipped = 0
    new_rows = []
    for row in generated_rows:
        if generated_row_key(row) in seen:
            skipped += 1
            continue
        max_line += 1
        new_rows.append(
            [
                max_line,
                row.driver_initials,
                row.po,
                "DR",
                row.move_date,
                row.machine,
                row.meter,
                row.description,
                row.origin,
                row.destination,
                row.order_number,
                "",
                "",
                "",
                "",
                "",
            ]
        )
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        for row in preserved + new_rows:
            writer.writerow(row)
    return len(new_rows), skipped


def write_lookup_csv(
    path: Path, lookup: dict[str, dict[str, str]], frequencies: Counter[str]
) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["serial", "description", "meter", "historical_matches"])
        for serial in sorted(lookup):
            writer.writerow(
                [
                    serial,
                    lookup[serial].get("description", ""),
                    lookup[serial].get("meter", "N/A"),
                    frequencies.get(serial, 0),
                ]
            )


def write_unresolved_csv(path: Path, unresolved: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "order_number",
                "serial",
                "source_column",
                "issue",
                "suggested_description",
            ],
        )
        writer.writeheader()
        writer.writerows(unresolved)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build a Ryan-format report from Axon exports."
    )
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--append-to")
    parser.add_argument("--append-output")
    parser.add_argument("--only-new-orders", action="store_true")
    parser.add_argument("--order-master")
    parser.add_argument("--new-ryan")
    parser.add_argument("--historical-ryan")
    parser.add_argument("--distribution-pdf",
        help="Path to Ryan's RIC EM Distribution Report PDF. If omitted, "
             "auto-discovered as the newest matching PDF in --input-dir.")
    parser.add_argument("--state-dir")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    if args.state_dir:
        state_dir = Path(args.state_dir)
    else:
        state_dir = Path(__file__).resolve().parents[1] / "state"
    state_dir.mkdir(parents=True, exist_ok=True)

    historical_path = (
        Path(args.historical_ryan)
        if args.historical_ryan
        else input_dir / "2026 RYAN MOVES.csv"
    )
    historical_lookup, historical_freq = parse_historical_ryan(historical_path)
    order_crosswalk = parse_historical_orders(historical_path)

    serial_lookup = build_serial_lookup(
        parse_overrides(state_dir / "generated_serial_lookup.csv"),
        historical_lookup,
        parse_overrides(state_dir / "serial_overrides.csv"),
    )

    # Distribution PDF: explicit path > auto-discovered in input_dir > none.
    # Used as fallback for equipment description on serials missing from history.
    distribution_pdf_path: Path | None = None
    if args.distribution_pdf:
        distribution_pdf_path = Path(args.distribution_pdf)
    else:
        distribution_pdf_path = discover_distribution_pdf(input_dir)
    distribution_lookup = (
        parse_distribution_pdf(distribution_pdf_path)
        if distribution_pdf_path
        else {}
    )
    if distribution_pdf_path:
        print(
            f"[INFO] Distribution PDF: {distribution_pdf_path.name} "
            f"({len(distribution_lookup)} assets mapped)"
        )

    # Primary path (new): Order Master Detail + Summary presets.
    # Fallback (legacy): Order Master Report (single export) + New RYAN.
    detail_path, summary_path = discover_order_master_pair(input_dir)
    if args.order_master:
        # Explicit path overrides discovery; assume it's a Summary export
        # if its content matches; else fall back to legacy parser.
        explicit = Path(args.order_master)
        kind = classify_order_master_csv(explicit)
        if kind == "summary":
            summary_path = explicit
        else:
            detail_path = explicit

    if summary_path is not None:
        # NEW path: build straight from the Summary report (Shipper/Consignee
        # already include the job#, Customer Job # is the destination fallback,
        # serials #1-4 expand to multiple rows in-function).
        drivers: dict[str, str] = {}
        if detail_path is not None:
            drivers = parse_order_master_detail_drivers(detail_path)
        print(f"[INFO] Using Summary path: {summary_path.name}"
              + (f" + Detail: {detail_path.name}" if detail_path else " (no Detail — Truck column will be blank)"))
        generated_rows, unresolved = parse_order_master_summary(
            summary_path,
            drivers=drivers,
            serial_lookup=serial_lookup,
            distribution_lookup=distribution_lookup,
        )
    else:
        # LEGACY path: Detail + New RYAN
        order_master_path = (
            Path(args.order_master)
            if args.order_master
            else discover_latest_file(input_dir, ORDER_MASTER_NAME)
        )
        new_ryan_path = (
            Path(args.new_ryan)
            if args.new_ryan
            else discover_latest_file(input_dir, NEW_RYAN_NAME)
        )
        order_master = parse_order_master(order_master_path)
        new_rows = parse_new_ryan(new_ryan_path)
        generated_rows, unresolved = collect_generated_rows(
            new_rows, order_master, serial_lookup,
            order_crosswalk, distribution_lookup,
        )

    skipped_existing_orders = 0
    if args.append_to and args.only_new_orders:
        generated_rows, skipped_existing_orders = filter_new_orders(
            generated_rows, Path(args.append_to)
        )
        unresolved = filter_unresolved_new_orders(unresolved, Path(args.append_to))

    generated_rows.sort(key=lambda r: (r.move_date, r.order_number, r.machine))
    write_target_csv(Path(args.output), generated_rows)
    write_lookup_csv(
        state_dir / "generated_serial_lookup.csv", serial_lookup, historical_freq
    )
    write_unresolved_csv(state_dir / "unresolved_serials.csv", unresolved)

    print(f"Generated rows: {len(generated_rows)}")
    print(f"Unresolved rows: {len(unresolved)}")
    if skipped_existing_orders:
        print(f"Skipped rows from existing orders: {skipped_existing_orders}")

    if args.append_to and args.append_output:
        appended, skipped = write_append_csv(
            Path(args.append_output), Path(args.append_to), generated_rows
        )
        print(f"Appended rows: {appended}")
        print(f"Skipped existing rows: {skipped}")


if __name__ == "__main__":
    main()
