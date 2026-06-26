"""Append generated Ryan rows into 2026 RYAN MOVES.xlsx with carry-over.

Behavior::

    1. Find the latest printable section in the '2026' sheet. If it has
       open template-padding rows (line-numbered rows with no order#),
       fill new rows into those slots first.
    2. When the carry-over fill reaches 25 rows, the section is "complete"
       and is recorded in the run summary so the caller can trigger an
       email draft for that page.
    3. Overflow rows that don't fit get written as additional new sections,
       wrapping at 25 lines each. Each new section gets a page-footer + banner
       + contact preamble (3 rows, matching the recent-section short style),
       then 25 data rows where unfilled trailing rows are template-padded with
       only their line number — so the NEXT run's carry-over works the same
       way.

The append always preserves cell formatting by copying styles from the
existing section template rows in the workbook (page-footer, banner, contact,
data). Datetime values are written as real datetimes with the '[$-409]d-mmm;@'
format, so dates print as '15-Apr' etc. Dedupe key is the tuple
(order#, machine#, date_iso, from, to) compared against every data row in
the '2026' sheet.

Usage::

    python execution/append_to_xlsx.py \\
        --xlsx "/path/to/2026 RYAN MOVES.xlsx" \\
        --from-csv /path/to/generated-ryan-report-new-only.csv \\
        --summary-out /path/to/completion-summary.json   # optional

The --summary-out option writes a JSON file describing exactly what happened
(rows appended, sections completed, page numbers, archive paths) — the daily
wrapper uses it to drive the email-draft step.
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import json
from copy import copy as _copy_style_obj
from pathlib import Path
from typing import Iterable

# openpyxl optionally imports numpy and, at IMPORT time, builds a tuple
# referencing numpy.short/ushort/etc. A PyInstaller-frozen numpy can be
# importable but INCOMPLETE (missing those attributes) -- that import then
# raises AttributeError and crashes the build ('module numpy has no attribute
# short'). We use numpy nowhere, so force its import to fail cleanly so openpyxl
# sets NUMPY=False and takes its pure-Python path. Unconditional override (not
# setdefault) so it beats a numpy PyInstaller may have pre-imported.
import sys as _sys
_sys.modules["numpy"] = None

from openpyxl import load_workbook
from openpyxl.cell import Cell

CONTACT_TEXT = "Catom Trucking 630-208-8442"
BANNER_TEXT = "RYAN INCORPORATED CENTRAL   EQUIPMENT MOVES"
EMAIL_TEXT = "equipment@ryancentral.com"

# Existing-section template rows in the '2026' sheet (used as style sources).
TPL_PAGE = 28
TPL_BANNER = 29
TPL_CONTACT = 30
TPL_DATA = 33

# Hard cap on data rows per printable section. Existing sections in the
# workbook print exactly 25 lines per page; once a section fills, the workflow
# auto-rolls to the next one. Match that here.
MAX_LINES_PER_SECTION = 25

# Number of columns to copy formatting for (matches existing sheet width).
NUM_COLS = 16


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _copy_style(src: Cell, dst: Cell) -> None:
    """Copy visual style from src cell to dst cell (font/fill/border/etc)."""
    if not src.has_style:
        return
    dst.font = _copy_style_obj(src.font)
    dst.fill = _copy_style_obj(src.fill)
    dst.border = _copy_style_obj(src.border)
    dst.alignment = _copy_style_obj(src.alignment)
    dst.number_format = src.number_format
    dst.protection = _copy_style_obj(src.protection)


def _coerce_int(value: str) -> int | str:
    """Return int(value) when it parses cleanly, else the trimmed string."""
    s = (value or "").strip()
    return int(s) if s.isdigit() else s


# ---------------------------------------------------------------------------
# Section discovery (carry-over support)
# ---------------------------------------------------------------------------

def _next_page_number(ws) -> int:
    """Return the next 'PAGE: NN OF 2026' number to use."""
    max_n = 0
    for row in ws.iter_rows(values_only=True):
        cell = row[7] if len(row) > 7 else None
        if isinstance(cell, str) and "PAGE:" in cell:
            try:
                token = cell.split("PAGE:", 1)[1].split("OF", 1)[0].strip()
                max_n = max(max_n, int(token))
            except (IndexError, ValueError):
                continue
    return max_n + 1


def _find_open_section(ws) -> dict | None:
    """Find the latest section's open template-padding rows, if any.

    A section is "open" when its trailing rows have line numbers in column A
    but no order # in column K — those are the template-padded rows ready to
    receive new entries. Returns a dict with keys::

        page_number   (int or None for the very first unbannered section)
        first_data_row    row index of line 1 of this section
        first_open_row    row index where the next entry should land
        next_line_no      line number to assign to the first new entry
        slots_remaining   how many template-padded rows are still empty

    Returns None when there is no open section (e.g. brand-new workbook).
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None

    # Latest banner row (1-indexed). If absent, treat the whole sheet as a
    # section starting at row 3 (after the two original column-header rows).
    last_banner_idx = None
    for i, row in enumerate(rows, start=1):
        if row and len(row) > 0 and row[0] == BANNER_TEXT:
            last_banner_idx = i
    if last_banner_idx is None:
        # Fallback: walk rows 3..max looking for the section's data range.
        first_data_row = 3
        page_number = None
    else:
        # Page-footer row sits one row above the banner; data rows start a
        # few rows below it. Sections in the workbook have either a 5-row
        # preamble (page-footer, banner, contact, hdr1, hdr2) or a 3-row
        # preamble (page-footer, banner, contact). Detect which.
        page_number = _read_page_number_at(rows, last_banner_idx - 1)
        first_data_row = _detect_first_data_row(rows, last_banner_idx)

    # Walk from first_data_row down, counting populated lines and finding the
    # first open template-padded row.
    populated = 0
    first_open_row = None
    next_line_no = 1
    last_line_no = 0
    for row_idx in range(first_data_row, len(rows) + 1):
        row = rows[row_idx - 1]
        line = row[0] if len(row) > 0 else None
        order_no = row[10] if len(row) > 10 else None
        if not isinstance(line, int):
            # Either we've walked off the end of the section, or hit the
            # padding ran out and we're now in some other content. Stop.
            break
        last_line_no = line
        if order_no in (None, ""):
            # Open template-padded row. Capture the first such row.
            if first_open_row is None:
                first_open_row = row_idx
                next_line_no = line
        else:
            populated += 1
            # If we'd previously found an open row and now see a populated
            # one again, the open row was a gap — don't carry over there.
            # In practice this doesn't happen in this workbook.
            if first_open_row is not None and row_idx <= first_open_row:
                first_open_row = None

    if first_open_row is None:
        # No template-padded slots — the section is effectively closed for
        # carry-over (or had no padding to begin with).
        return None

    slots_remaining = max(0, MAX_LINES_PER_SECTION - populated)
    return {
        "page_number": page_number,
        "first_data_row": first_data_row,
        "first_open_row": first_open_row,
        "next_line_no": next_line_no,
        "slots_remaining": slots_remaining,
        "populated_before": populated,
        "last_line_no": last_line_no,
    }


def _read_page_number_at(rows: list, row_idx: int) -> int | None:
    """Parse 'PAGE: NN OF 2026' out of column H of the given 1-indexed row."""
    if row_idx < 1 or row_idx > len(rows):
        return None
    row = rows[row_idx - 1]
    cell = row[7] if len(row) > 7 else None
    if isinstance(cell, str) and "PAGE:" in cell:
        try:
            token = cell.split("PAGE:", 1)[1].split("OF", 1)[0].strip()
            return int(token)
        except (IndexError, ValueError):
            return None
    return None


def _detect_first_data_row(rows: list, banner_row: int) -> int:
    """Return the 1-indexed row where data lines start for a given banner.

    Sections in this workbook either have:
      - a 5-row preamble: page-footer, banner, contact, hdr1, hdr2 -> data at banner+4
      - a 3-row preamble: page-footer, banner, contact -> data at banner+2

    Detect by checking whether the row immediately after contact has 'Truck'
    in column B (header row) or an integer in column A (data row).
    """
    # Try short-style first: data at banner + 2 (banner, contact, then data)
    candidate = banner_row + 2
    if candidate <= len(rows):
        row = rows[candidate - 1]
        if len(row) > 0 and isinstance(row[0], int):
            return candidate
    # Otherwise, assume the older 5-row style with headers at banner+2,3
    return banner_row + 4


# ---------------------------------------------------------------------------
# Existing-data dedupe
# ---------------------------------------------------------------------------

def _existing_keys(ws) -> set[tuple[str, str, str, str, str]]:
    """Return set of (order#, machine#, date_iso, from, to) for dedupe."""
    keys: set[tuple[str, str, str, str, str]] = set()
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 11:
            continue
        line, _truck, _po, _by, date, machine, _meter, _desc, frm, to, order_no = row[:11]
        if not isinstance(line, int):
            continue
        if isinstance(date, _dt.datetime):
            date_iso = date.strftime("%Y-%m-%d")
        elif isinstance(date, _dt.date):
            date_iso = date.strftime("%Y-%m-%d")
        else:
            parsed = _parse_d_mmm(str(date or "").strip())
            date_iso = parsed.strftime("%Y-%m-%d") if parsed else str(date or "").strip()
        keys.add(
            (
                str(order_no or "").strip(),
                str(machine or "").strip(),
                date_iso,
                str(frm or "").strip(),
                str(to or "").strip(),
            )
        )
    return keys


def _parse_d_mmm(s: str, today: _dt.date | None = None) -> _dt.datetime | None:
    """Parse a move-date string into a datetime.

    Accepts BOTH the 'd-mmm' form produced by build_ryan_report.format_move_date
    (e.g. '16-Jun') AND full dates like '06/16/2026' / '2026-06-16' that flow
    through from the source reports. The full-date case used to fail here, which
    silently blanked the Date column in the workbook — the #1 'missing date' bug.
    """
    s = (s or "").strip()
    if not s:
        return None
    today = today or _dt.date.today()
    # Full-date formats already carry their year — take them verbatim.
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%-m/%-d/%Y"):
        try:
            return _dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    # 'd-mmm' forms have no year — assume the current year, rolling back if that
    # lands more than 60 days in the future.
    parsed = None
    for fmt in ("%d-%b", "%-d-%b"):
        try:
            parsed = _dt.datetime.strptime(s, fmt).replace(year=today.year)
            break
        except ValueError:
            continue
    if parsed is None:
        return None
    if parsed.date() > today + _dt.timedelta(days=60):
        parsed = parsed.replace(year=today.year - 1)
    return parsed


def _read_generated_csv(csv_path: Path) -> list[dict[str, str]]:
    """Read the new-only CSV produced by build_ryan_report.py."""
    rows: list[dict[str, str]] = []
    with csv_path.open("r", encoding="utf-8", newline="") as h:
        for raw in csv.reader(h):
            if len(raw) < 11:
                continue
            line = raw[0].strip()
            if not line.isdigit():
                continue
            rows.append(
                {
                    "line": line,
                    "truck": raw[1].strip(),
                    "po": raw[2].strip(),
                    "by": raw[3].strip() or "DR",
                    "date_str": raw[4].strip(),
                    "machine": raw[5].strip(),
                    "meter": raw[6].strip() or "N/A",
                    "description": raw[7].strip(),
                    "from": raw[8].strip(),
                    "to": raw[9].strip(),
                    "order": raw[10].strip(),
                }
            )
    return rows


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

def _write_data_row(
    ws,
    row_idx: int,
    line_no: int,
    gen_row: dict[str, str],
    date_dt: _dt.datetime | None,
) -> None:
    """Style + value-fill a data row (overwriting whatever was there)."""
    for col in range(1, NUM_COLS + 1):
        _copy_style(ws.cell(row=TPL_DATA, column=col), ws.cell(row=row_idx, column=col))
    ws.cell(row=row_idx, column=1).value = line_no
    ws.cell(row=row_idx, column=2).value = gen_row["truck"]
    ws.cell(row=row_idx, column=3).value = _coerce_int(gen_row["po"])
    ws.cell(row=row_idx, column=4).value = gen_row["by"]
    ws.cell(row=row_idx, column=5).value = date_dt
    ws.cell(row=row_idx, column=6).value = gen_row["machine"]
    ws.cell(row=row_idx, column=7).value = gen_row["meter"]
    ws.cell(row=row_idx, column=8).value = gen_row["description"]
    ws.cell(row=row_idx, column=9).value = gen_row["from"]
    ws.cell(row=row_idx, column=10).value = gen_row["to"]
    ws.cell(row=row_idx, column=11).value = _coerce_int(gen_row["order"])
    ws.row_dimensions[row_idx].height = (
        ws.row_dimensions[TPL_DATA].height or 15.75
    )


def _write_template_padding_row(ws, row_idx: int, line_no: int) -> None:
    """Write a line-numbered template-padding row (col A only, rest empty)."""
    for col in range(1, NUM_COLS + 1):
        _copy_style(ws.cell(row=TPL_DATA, column=col), ws.cell(row=row_idx, column=col))
        if col != 1:
            ws.cell(row=row_idx, column=col).value = None
    ws.cell(row=row_idx, column=1).value = line_no
    ws.row_dimensions[row_idx].height = (
        ws.row_dimensions[TPL_DATA].height or 15.75
    )


def _write_new_section_preamble(ws, start_row: int, page_number: int) -> int:
    """Write a 3-row preamble (page-footer + banner + contact). Returns the
    row immediately after the preamble (where data lines should begin).

    Defensive: explicitly nulls every column in each preamble row before
    writing the small handful of values we actually want, so this can never
    inherit garbage from a recycled cell. The historical xlsx had pages
    with stale paste-typo strings (e.g. an email-and-job# concat) leaking
    into col I of every page-footer row; this guarantees future writes
    won't reintroduce that.
    """
    preamble = [
        (TPL_PAGE,    [(1, EMAIL_TEXT), (8, f"PAGE:  {page_number:02d}  OF   2026")]),
        (TPL_BANNER,  [(1, BANNER_TEXT)]),
        (TPL_CONTACT, [(1, CONTACT_TEXT)]),
    ]
    cur = start_row
    for tpl_row, values in preamble:
        for col in range(1, NUM_COLS + 1):
            cell = ws.cell(row=cur, column=col)
            _copy_style(ws.cell(row=tpl_row, column=col), cell)
            cell.value = None  # clear before assigning, no garbage carry-over
        for col, val in values:
            ws.cell(row=cur, column=col).value = val
        ws.row_dimensions[cur].height = (
            ws.row_dimensions[tpl_row].height or 15.75
        )
        cur += 1
    return cur


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def append_section(
    xlsx_path: Path,
    generated_rows: Iterable[dict[str, str]],
    sheet_name: str = "2026",
    today: _dt.date | None = None,
) -> dict:
    """Carry-over append. Returns a structured run summary."""
    rows_list = list(generated_rows)
    summary: dict = {
        "appended": 0,
        "skipped_dupe": 0,
        "carry_over_filled": 0,
        "new_sections_written": 0,
        "completed_sections": [],
        "first_touched_row": None,
    }

    if not rows_list:
        return summary

    lock = xlsx_path.parent / f"~${xlsx_path.name}"
    if lock.exists():
        raise RuntimeError(
            f"Excel has '{xlsx_path.name}' open ({lock.name} present). "
            f"Close the file in Excel and re-run."
        )
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise RuntimeError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    # Dedupe before doing anything destructive.
    existing = _existing_keys(ws)
    fresh: list[tuple[dict[str, str], _dt.datetime | None]] = []
    skipped = 0
    for r in rows_list:
        date_dt = _parse_d_mmm(r["date_str"], today=today)
        date_iso = date_dt.strftime("%Y-%m-%d") if date_dt else ""
        key = (r["order"], r["machine"], date_iso, r["from"], r["to"])
        if key in existing:
            skipped += 1
            continue
        fresh.append((r, date_dt))
    summary["skipped_dupe"] = skipped

    # Sort chronologically by parsed datetime so each printable section has
    # its dates in order (and pages don't mix Feb with April). Rows with an
    # unparseable date go last, ordered by their original CSV position so we
    # don't reshuffle them randomly. Tie-break on order# then machine# to
    # match the build script's secondary sort intent.
    _UNDATED = _dt.datetime.max

    def _sort_key(item):
        gen_row, dt = item
        return (
            dt or _UNDATED,
            str(gen_row.get("order") or ""),
            str(gen_row.get("machine") or ""),
        )

    fresh.sort(key=_sort_key)

    if not fresh:
        wb.close()
        return summary

    # ---- Step 1: carry-over fill the latest open section ----
    open_section = _find_open_section(ws)
    queue = list(fresh)
    if open_section and open_section["slots_remaining"] > 0:
        n_to_fill = min(len(queue), open_section["slots_remaining"])
        cur_row = open_section["first_open_row"]
        cur_line = open_section["next_line_no"]
        for i in range(n_to_fill):
            gen_row, date_dt = queue[i]
            _write_data_row(ws, cur_row, cur_line, gen_row, date_dt)
            if summary["first_touched_row"] is None:
                summary["first_touched_row"] = cur_row
            cur_row += 1
            cur_line += 1
        summary["carry_over_filled"] = n_to_fill
        summary["appended"] += n_to_fill
        # Did this section just hit 25?
        new_total = open_section["populated_before"] + n_to_fill
        if new_total >= MAX_LINES_PER_SECTION and open_section["page_number"] is not None:
            summary["completed_sections"].append(
                {
                    "page_number": open_section["page_number"],
                    "completion_type": "carry_over",
                    "rows_added_this_run": n_to_fill,
                    "section_first_row": open_section["first_data_row"],
                }
            )
        queue = queue[n_to_fill:]

    # ---- Step 2: overflow into new section(s), 25 rows each ----
    while queue:
        chunk = queue[:MAX_LINES_PER_SECTION]
        queue = queue[MAX_LINES_PER_SECTION:]
        page_n = _next_page_number(ws)

        # Place the new section's preamble after whatever is currently the
        # last row in the sheet.
        preamble_start = ws.max_row + 1
        data_start = _write_new_section_preamble(ws, preamble_start, page_n)
        if summary["first_touched_row"] is None:
            summary["first_touched_row"] = preamble_start

        for line_no, (gen_row, date_dt) in enumerate(chunk, start=1):
            _write_data_row(ws, data_start + line_no - 1, line_no, gen_row, date_dt)

        # Fill the rest of the 25 data slots with template-padding rows so
        # the next run's carry-over logic finds open slots.
        for line_no in range(len(chunk) + 1, MAX_LINES_PER_SECTION + 1):
            _write_template_padding_row(ws, data_start + line_no - 1, line_no)

        summary["new_sections_written"] += 1
        summary["appended"] += len(chunk)
        if len(chunk) >= MAX_LINES_PER_SECTION:
            summary["completed_sections"].append(
                {
                    "page_number": page_n,
                    "completion_type": "new",
                    "rows_added_this_run": len(chunk),
                    "section_first_row": data_start,
                }
            )

    wb.save(xlsx_path)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Append generated Ryan rows into the canonical 2026 RYAN MOVES.xlsx, "
            "filling open template slots first (carry-over) and overflowing into "
            "new printable sections wrapping at 25 rows each."
        )
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        help="Path to the workbook (e.g. '~/Downloads/2026 RYAN MOVES.xlsx').",
    )
    parser.add_argument(
        "--from-csv",
        required=True,
        help="CSV produced by build_ryan_report.py --output (the new-only CSV).",
    )
    parser.add_argument(
        "--sheet",
        default="2026",
        help="Sheet to append to (default: '2026').",
    )
    parser.add_argument(
        "--summary-out",
        default="",
        help="Optional path to write a JSON run summary (used by the daily wrapper).",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser()
    csv_path = Path(args.from_csv).expanduser()

    rows = _read_generated_csv(csv_path)
    if not rows:
        print(f"No generated rows in {csv_path}; nothing to append.")
        if args.summary_out:
            Path(args.summary_out).expanduser().write_text(
                json.dumps(
                    {
                        "appended": 0,
                        "skipped_dupe": 0,
                        "carry_over_filled": 0,
                        "new_sections_written": 0,
                        "completed_sections": [],
                    },
                    indent=2,
                )
            )
        return

    summary = append_section(xlsx_path, rows, sheet_name=args.sheet)
    print(
        f"Appended: {summary['appended']}  "
        f"Skipped (duplicates): {summary['skipped_dupe']}  "
        f"Carry-over filled: {summary['carry_over_filled']}  "
        f"New sections: {summary['new_sections_written']}  "
        f"Completed sections this run: {len(summary['completed_sections'])}"
    )
    for cs in summary["completed_sections"]:
        print(
            f"  -> Section completed: page {cs['page_number']:02d}  "
            f"({cs['completion_type']}, +{cs['rows_added_this_run']} rows this run)"
        )

    if args.summary_out:
        Path(args.summary_out).expanduser().write_text(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
